
import sqlite3
import openpyxl

wb = openpyxl.load_workbook('Pitching20.xlsx')
wbsheet = wb["Pitching"]
# Let's read from the excel file in a script and make sure things look correct
vals = ''
for i in range(2, 11): # rows - 2 is first record
    for j in range(1,31): # columns
        print(i, j, wbsheet.cell(row=i, column=j).value)
        vals = vals + '"' + str(wbsheet.cell(row=i, column=j).value) + '",'
    # add right paren
    vals = vals + ')'
    # replace "None" with NULL
    vals.replace('None','NULL')
    # replace last comma
    vals.replace(',)',')')
    print('insert into pitching values (' + vals)
    vals = ''
